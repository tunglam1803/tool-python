import asyncio
import os
import random
import string
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font

# Configuration
SHEET_URL = "https://docs.google.com/spreadsheets/d/1LuNG22iALJ9aXtPlubqtiV9UhHr12yDOU-gDv3sLlh4/edit#gid=1636014016"
GCP_WELCOME_URL = "https://console.cloud.google.com/home/dashboard"
GCP_PROJECT_CREATE_URL = "https://console.cloud.google.com/projectcreate"
PORTAL_URL = "https://taphoa.qzz.io/customer-portal"
MGMT_OAUTH_URL = "http://171.244.130.59:8317/management.html#/oauth"
XLSX_FILENAME = "temp_accounts.xlsx"

# Selectors (GCP Login)
GOOGLE_ID = "input#identifierId"
GOOGLE_NEXT = "xpath=//button[.//span[text()='Next' or text()='Tiếp theo' or text()='Tiếp']]"
AZURE_ID = "input#i0116"
AZURE_PASS = "input#i0118"
AZURE_NEXT = "input#idSIButton9"

# Portal Selectors
PORTAL_API_KEY = "input[placeholder*='ek_']"
PORTAL_EMAIL = "input[placeholder='user@example.com']"
PORTAL_INVOKE = "button:has-text('Invoke Service')"

# Management Selectors (Gemini CLI OAuth section)
MGMT_SECTION_SELECTOR = "div.card:has-text('Gemini CLI OAuth')"
MGMT_PROJECT_ID_INPUT = "input[placeholder*='auto-select first available project']"
MGMT_LOGIN_BTN = "button:has-text('Login')"
MGMT_CALLBACK_INPUT = "input[placeholder*='auth/callback?code=']"
MGMT_CALLBACK_SUBMIT = "button:has-text('Submit Callback URL')"

def generate_project_name():
    """Generates a random project name."""
    return "project-" + "".join(random.choices(string.ascii_lowercase + string.digits, k=6))

async def get_accounts_from_excel(file_path):
    print(f"[*] Reading data from local file: {file_path}")
    accounts = []
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            email = str(row[1].value).strip() if row[1].value else ""
            password = str(row[2].value).strip() if row[2].value else ""
            api_key = str(row[3].value).strip() if row[3].value else ""
            if not email or "@" not in email: continue
            
            # Check if bold (Done) or has any color (Error)
            font = row[1].font
            is_bold = font.bold if font else False
            # Check if it has any RGB color (indicating an error was already marked)
            has_color = (font and font.color and font.color.type == 'rgb' and font.color.rgb != "00000000")
            
            if not is_bold and not has_color:
                accounts.append({"row_idx": row_idx, "email": email, "password": password, "api_key": api_key})
        wb.close()
        return accounts
    except Exception as e:
        print(f"[-] Error reading Excel: {e}")
        return []

def mark_account_as_bold_in_excel(file_path, row_idx):
    """Updates the local Excel file to mark the email as bold."""
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        sheet.cell(row=row_idx, column=2).font = Font(bold=True)
        wb.save(file_path)
        wb.close()
        print(f"[+] Local Excel updated: Row {row_idx} marked as bold.")
    except Exception as e:
        print(f"[-] Error updating Local Excel: {e}")

def mark_account_as_error_in_excel(file_path, row_idx, error_msg, color="FF0000"):
    """Marks the email with a specific color and writes the error message to column E."""
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        sheet.cell(row=row_idx, column=2).font = Font(color=color)
        sheet.cell(row=row_idx, column=5).value = error_msg
        wb.save(file_path)
        wb.close()
        print(f"[!] Local Excel updated: Row {row_idx} marked as {error_msg}.")
    except Exception as e:
        print(f"[-] Error marking error in Excel: {e}")

def is_account_done_in_excel(file_path, row_idx):
    """Re-checks if an account is already marked as bold before processing."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active
        cell = sheet.cell(row=row_idx, column=2)
        is_bold = cell.font.bold if cell.font else False
        wb.close()
        return is_bold
    except:
        return False

async def sync_status_to_google_sheet(page, email, status_msg=None, color_name=None, is_bold=False):
    """Syncs status, color, and bold state to Google Sheets via keyboard shortcuts."""
    print(f"[*] Syncing status to Google Sheet for {email}...")
    try:
        await page.bring_to_front()
        await asyncio.sleep(1)
        # Search for email (Column B)
        await page.keyboard.press("Control+f")
        await asyncio.sleep(1)
        await page.keyboard.type(email, delay=100)
        await page.keyboard.press("Enter")
        await asyncio.sleep(1)
        await page.keyboard.press("Enter") # Select the cell
        await asyncio.sleep(1)
        await page.keyboard.press("Escape") # Close search box
        await asyncio.sleep(1)

        # 1. PRIORITY: Bold the row first (so we don't lose focus)
        if is_bold:
            await page.keyboard.down("Shift")
            await page.keyboard.press("Space")
            await page.keyboard.up("Shift")
            await asyncio.sleep(1)
            await page.keyboard.press("Control+b")
            await asyncio.sleep(1)

        # 2. Set Font Color if needed
        if color_name:
            await page.keyboard.press("Alt+/")
            await asyncio.sleep(1)
            await page.keyboard.type(f"text color {color_name}")
            await asyncio.sleep(1)
            await page.keyboard.press("Enter")
            await asyncio.sleep(1)

        # 3. Write status msg to Column E (3 steps right from Column B)
        if status_msg:
            # Move to Column E
            for _ in range(3): await page.keyboard.press("ArrowRight")
            await page.keyboard.press("Enter")
            await page.keyboard.type(status_msg, delay=50)
            await page.keyboard.press("Enter")
            await asyncio.sleep(1)
            
    except Exception as e:
        print(f"[!] Error syncing to Google Sheet: {e}")

async def handle_portal_verification(page, account):
    """Portal verification (Optional)"""
    print(f"[*] Portal verifying: {account['email']}")
    try:
        portal_page = await page.context.new_page()
        await portal_page.goto(PORTAL_URL, timeout=30000)
        await portal_page.fill(PORTAL_API_KEY, account['api_key'])
        await portal_page.fill(PORTAL_EMAIL, account['email'])
        await portal_page.click(PORTAL_INVOKE)
        await asyncio.sleep(3)
        await portal_page.close()
        return True
    except Exception as e:
        print(f"[!] Portal skipped/error: {e}")
        return False

async def handle_gcp_onboarding(page):
    """Aggressively handles GCP onboarding dialogs and speedbumps."""
    print("[*] Clearing GCP onboarding dialog...")
    try:
        # 1. Handle "I understand" Speedbumps (Aggressive Loop)
        max_speedbump_attempts = 5
        for _ in range(max_speedbump_attempts):
            if "speedbump" not in page.url and "accounts.google.com" not in page.url:
                break
                
            print(f"[*] Speedbump detected at {page.url}. Clearing...")
            try:
                # Scroll to bottom
                await page.evaluate("window.scrollTo(0, 5000)")
                await asyncio.sleep(1)
                
                # Check common buttons
                selectors = [
                    "button:has-text('I understand')",
                    "button:has-text('Tôi đã hiểu')",
                    "button:has-text('Confirm')",
                    "button:has-text('Xác nhận')",
                    "button:has-text('Agree')",
                    "button:has-text('Đồng ý')",
                    "button:has-text('Done')",
                    "button:has-text('Xong')",
                    "text='I understand'",
                    "text='Confirm'",
                    "text='Done'"
                ]
                
                clicked = False
                for s in selectors:
                    btn = page.locator(s).first
                    if await btn.is_visible(timeout=2000):
                        print(f"[*] Found button with '{s}'. Clicking...")
                        await btn.click(force=True)
                        await asyncio.sleep(4)
                        clicked = True
                        break
                
                if not clicked:
                    print("[!] No speedbump button detected but page is still active. Waiting...")
                    await asyncio.sleep(3)
            except: 
                break
        
        # 2. Handle Terms of Service Checkbox (Inside GCP console)
        if "console.cloud.google.com" in page.url:
            import re
            try:
                checkbox = page.locator("input[type='checkbox']").first
                if await checkbox.is_visible(timeout=3000):
                    print("[*] Ticking terms checkbox...")
                    await checkbox.click(force=True)
                    await asyncio.sleep(1)
                else:
                    terms_el = page.locator("mat-checkbox, label").filter(has_text=re.compile(r"I agree to the Google Cloud Platform Terms of Service|Tôi đồng ý với Điều khoản dịch vụ", re.I)).first
                    if await terms_el.is_visible(timeout=1000):
                        print("[*] Clicking terms text...")
                        await terms_el.click(force=True)
                        await asyncio.sleep(1)
                
                # Agree Button
                agree_btn = page.locator("button:has-text('Agree and continue'), button:has-text('Đồng ý và tiếp tục')").first
                if await agree_btn.is_visible(timeout=3000):
                    print("[*] Clicking Agree & Continue...")
                    await agree_btn.click(force=True)
                    await asyncio.sleep(2)
            except: pass

        # 3. Close feedback popups
        try:
            close_btn = page.locator("button[aria-label='Close'], button[aria-label='Đóng'], .cfc-feedback-close-button").first
            if await close_btn.is_visible(timeout=1000):
                await close_btn.click(force=True)
        except: pass

    except Exception as e:
        print(f"[!] Onboarding Handler Error: {e}")

async def create_gcp_project(page):
    """Creates a project by clicking Create immediately."""
    for attempt in range(3):
        print(f"[*] Navigating to Project Creation (Attempt {attempt+1})...")
        try:
            await page.goto(GCP_PROJECT_CREATE_URL, timeout=60000)
            await asyncio.sleep(10)
            await handle_gcp_onboarding(page)
            
            project_id = await page.evaluate("""() => {
                const el = Array.from(document.querySelectorAll('span, div')).find(e => e.innerText.includes('Project ID:'));
                if (el) {
                    const match = el.innerText.match(/Project ID: ([a-z0-9-]+)/);
                    return match ? match[1].trim() : '';
                }
                return '';
            }""")
            
            if not project_id:
                print("[!] Could not capture ID. Retrying...")
                await handle_gcp_onboarding(page)
                continue
                
            print(f"[+] Captured Project ID: {project_id}")
            print("[*] Clicking CREATE button...")
            await page.click("button:has-text('Create'), button:has-text('Tạo')", force=True, timeout=15000)
            
            try:
                await page.wait_for_url("**/console.cloud.google.com/welcome**", timeout=30000)
                print("[+] Project created successfully.")
            except: pass
            
            return project_id
        except Exception as e:
            print(f"[-] Attempt {attempt+1} failed: {e}")
    return None

async def handle_gemini_oauth(gcp_page, project_id, account_email):
    """Handles Gemini OAuth sequence, including account selection in GCP tab."""
    print(f"[*] Starting Gemini OAuth for: {project_id}")
    try:
        # Open Management in same context
        mgmt_page = await gcp_page.context.new_page()
        await mgmt_page.goto(MGMT_OAUTH_URL)
        await asyncio.sleep(2)
        
        if "/login" in mgmt_page.url:
            print("[*] Management login required...")
            await mgmt_page.fill("input[type='password']", "Nkg@6688")
            await mgmt_page.keyboard.press("Enter")
            await asyncio.sleep(3)
            await mgmt_page.goto(MGMT_OAUTH_URL)

        # 2. Enter Project ID and click Login
        gemini_section = mgmt_page.locator(MGMT_SECTION_SELECTOR)
        await gemini_section.scroll_into_view_if_needed()
        await gemini_section.locator(MGMT_PROJECT_ID_INPUT).fill(project_id)
        await gemini_section.locator(MGMT_LOGIN_BTN).click()
        
        # 3. Wait for Authorization URL
        auth_url_locator = gemini_section.locator("div[class*='authUrlValue']").first
        await auth_url_locator.wait_for(state="visible", timeout=15000)
        auth_url = (await auth_url_locator.inner_text()).strip()
        print(f"[+] Captured Auth URL: {auth_url[:50]}...")

        # 4. Navigate GCP tab to Auth URL
        print("[*] Navigating GCP tab to Auth URL...")
        callback_url = None
        def catch_request(request):
            nonlocal callback_url
            if "localhost:8085" in request.url: callback_url = request.url

        gcp_page.on("request", catch_request)
        await gcp_page.bring_to_front()
        await gcp_page.goto(auth_url)
        
        # --- NEW: Select the account if "Choose an account" appears ---
        try:
            # Look for the email text anywhere to click it
            email_locator = gcp_page.locator(f"text={account_email}").first
            if await email_locator.is_visible(timeout=5000):
                print(f"[*] Selecting account: {account_email}")
                await email_locator.click(force=True)
                await asyncio.sleep(3)
        except: pass

        # --- NEW: Click "Sign in" if confirmation screen appears ---
        try:
            # The confirmation screen has a "Sign in" button
            signin_btn = gcp_page.locator("button:has-text('Sign in'), button:has-text('Đăng nhập')").first
            if await signin_btn.is_visible(timeout=5000):
                print("[*] Clicking Sign-in confirmation...")
                await signin_btn.click(force=True)
                await asyncio.sleep(3)
        except: pass
        # -------------------------------------------------------------

        # Handle "Allow" if it appears
        try: await gcp_page.click("button:has-text('Allow'), button:has-text('Cho phép')", timeout=10000)
        except: pass
        
        for _ in range(25):
            if callback_url: break
            await asyncio.sleep(1)
        
        if not callback_url:
            print("[-] Failed to capture Callback URL.")
            await mgmt_page.close()
            return False
            
        # 5. Submit Callback back in Management tab
        await mgmt_page.bring_to_front()
        await gemini_section.locator(MGMT_CALLBACK_INPUT).fill(callback_url)
        print("[*] Submitting Callback URL. Waiting (at least 20s) for success message...")
        await gemini_section.locator(MGMT_CALLBACK_SUBMIT).click()
        
        # 6. Success Verification
        # Increased timeout to 60s as it takes at least 20s to appear
        success_label = mgmt_page.get_by_text("Authentication successful!").first
        await success_label.wait_for(state="visible", timeout=60000)
        print(f"[+] Gemini OAuth SUCCESS for {account_email}!")
        print("[*] Waiting 10 seconds before moving to next account...")
        await asyncio.sleep(10)
        
        await mgmt_page.close()
        return True
    except Exception as e:
        print(f"[-] Gemini OAuth Error: {e}")
        return False

async def main():
    xlsx_path = os.path.join(os.getcwd(), XLSX_FILENAME)
    if not os.path.exists(xlsx_path): return

    accounts = await get_accounts_from_excel(xlsx_path)
    if not accounts: 
        print("[*] No unprocessed accounts found.")
        return

    # --- REVERSE ORDER FOR THIS FILE ---
    print("[*] REVERSE MODE: Processing from bottom up...")
    accounts.reverse()
    # ------------------------------------

    async with async_playwright() as p:
        browser_opt = {
            "channel": "msedge", 
            "headless": False,
            "ignore_default_args": ["--enable-automation"],
            "args": ["--disable-blink-features=AutomationControlled"]
        }
        
        # Google Sheets Context
        user_data_path = os.path.join(os.getcwd(), "browser_session_2")
        try: gcp_context = await p.chromium.launch_persistent_context(user_data_path, **browser_opt)
        except: gcp_context = await p.chromium.launch_persistent_context(user_data_path, channel="chrome", headless=False)
            
        sheet_page = await gcp_context.new_page()
        await sheet_page.goto(SHEET_URL)
        if "signin" in sheet_page.url:
            await sheet_page.wait_for_url(f"*{SHEET_URL}*", timeout=600000)

        async def ensure_sheet_page():
            nonlocal sheet_page
            if sheet_page.is_closed():
                print("[!] Sheet page closed. Reopening...")
                sheet_page = await gcp_context.new_page()
                await sheet_page.goto(SHEET_URL)
                await asyncio.sleep(2)
            return sheet_page

        # Reg Browser instance (shared for efficiency)
        reg_browser = await p.chromium.launch(**browser_opt)

        for account in accounts:
            # Re-check if already done (crucial for parallel runs)
            if is_account_done_in_excel(xlsx_path, account['row_idx']):
                print(f"[*] Skipping {account['email']} - done by other instance or already bold.")
                continue

            # Re-launch browser if it was closed or crashed
            if not reg_browser.is_connected():
                print("[!] Registration browser disconnected. Relaunching...")
                reg_browser = await p.chromium.launch(**browser_opt)
                
            reg_context = await reg_browser.new_context()
            reg_page = await reg_context.new_page()
            
            print(f"\n[>>>] WORKING ON: {account['email']}")
            try:
                await reg_page.goto(GCP_WELCOME_URL)
                await reg_page.wait_for_selector(GOOGLE_ID, timeout=2000) # Fast check if already logged in
            except: pass

            try:
                if "identifier" in reg_page.url:
                    await reg_page.fill(GOOGLE_ID, account['email']); await reg_page.click(GOOGLE_NEXT)
                    await reg_page.wait_for_selector(AZURE_ID, timeout=40000)
                    await reg_page.fill(AZURE_ID, account['email']); await reg_page.press(AZURE_ID, "Enter")
                    await reg_page.wait_for_selector(AZURE_PASS, timeout=30000)
                    await reg_page.fill(AZURE_PASS, account['password']); await reg_page.click(AZURE_NEXT)
                    
                    # Check for Specific Duo Device Management Portal Block
                    try:
                        if "devicemanagement.duosecurity.com" in reg_page.url:
                            print(f"[!] Duo Device Management detected for {account['email']}. Skipping...")
                            mark_account_as_error_in_excel(xlsx_path, account['row_idx'], "MFA/Duo Required", color="FFA500")
                            sp = await ensure_sheet_page()
                            await sync_status_to_google_sheet(sp, account['email'], status_msg="MFA Required", color_name="orange")
                            try: await reg_context.close()
                            except: pass
                            continue
                    except: pass

                    try: await reg_page.click("input#idSIButton9", timeout=5000)
                    except: pass
                
                await handle_portal_verification(reg_page, account)
                try:
                    confirm_btn = "button:has-text('Yes, this is my device'), button:has-text('Có')"
                    await reg_page.wait_for_selector(confirm_btn, timeout=15000)
                    await reg_page.click(confirm_btn)
                except: pass
                
                print("[*] Waiting for Google Cloud Console...")
                # Immediate check for Duo Portal redirect
                if "devicemanagement.duosecurity.com" in reg_page.url:
                    print(f"[!] Duo Device Management detected for {account['email']}. Skipping...")
                    mark_account_as_error_in_excel(xlsx_path, account['row_idx'], "MFA/Duo Required", color="FFA500")
                    sp = await ensure_sheet_page()
                    await sync_status_to_google_sheet(sp, account['email'], status_msg="MFA Required", color_name="orange")
                    try: await reg_context.close()
                    except: pass
                    continue

                # Proactively call onboarding handler
                for _ in range(3):
                    # Re-check Duo inside loop in case of slow redirect
                    if "devicemanagement.duosecurity.com" in reg_page.url:
                        break
                    await handle_gcp_onboarding(reg_page)
                    await asyncio.sleep(2)
                    # Break early if we see console indicators
                    if await reg_page.locator(".cfc-logo, input[aria-label*='Search']").first.is_visible():
                        break
                
                # One last check before the long wait
                if "devicemanagement.duosecurity.com" in reg_page.url:
                    print(f"[!] Duo Device Management detected for {account['email']}. Skipping...")
                    mark_account_as_error_in_excel(xlsx_path, account['row_idx'], "MFA/Duo Required", color="FFA500")
                    sp = await ensure_sheet_page()
                    await sync_status_to_google_sheet(sp, account['email'], status_msg="MFA Required", color_name="orange")
                    try: await reg_context.close()
                    except: pass
                    continue

                try: await reg_page.locator(".cfc-logo, input[aria-label*='Search']").first.wait_for(state="visible", timeout=20000)
                except: await reg_page.wait_for_url("**/console.cloud.google.com/**", timeout=10000)
                
                await asyncio.sleep(2)
                project_id = await create_gcp_project(reg_page)
                
                if project_id:
                    # OAuth NOW RUNS IN THE SAME BROWSER WINDOW
                    if await handle_gemini_oauth(reg_page, project_id, account['email']):
                        # SUCCESS: Mark in Sheet
                        print(f"[*] Marking {account['email']} as DONE on Google Sheet...")
                        sp = await ensure_sheet_page()
                        await sync_status_to_google_sheet(sp, account['email'], status_msg="Done", is_bold=True)
                        
                        # SUCCESS: Sync back to local Excel
                        mark_account_as_bold_in_excel(xlsx_path, account['row_idx'])
                        print(f"[Done] {account['email']}")

            except Exception as e:
                print(f"[-] Error during account processing: {e}")
                # Log general error to both files
                mark_account_as_error_in_excel(xlsx_path, account['row_idx'], f"Error: {str(e)[:50]}")
                try:
                    sp = await ensure_sheet_page()
                    await sync_status_to_google_sheet(sp, account['email'], status_msg="Error", color_name="red")
                except: pass
            
            # Close context but keep browser open
            try: await reg_context.close()
            except: pass

        print("\n[*] ALL TASKS FINISHED!")
        await gcp_context.close()
        await reg_browser.close()

if __name__ == "__main__":
    asyncio.run(main())
